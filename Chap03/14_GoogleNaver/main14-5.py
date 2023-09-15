from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb = load_workbook(r'.//이메일주소.xlsx', data_only=True)
load_ws = load_wb.active

for i in range(1, load_ws.max_row + 1):
    recv_email_value = load_ws.cell(i, 1).value
    print("성공:", recv_email_value)
    try:
        send_email = "0306jh@naver.com"
        send_pwd = "@victory5637"

        recv_email = recv_email_value

        smtp_name = "smtp.naver.com"
        smtp_port = 587

        msg = MIMEMultipart()

        msg['Subject'] = '네이버에서 구글로 보내지는 엑셀메일정보 입니다.'
        msg['From'] = send_email
        msg['To'] = recv_email

        text = """
        네이버에서 구글로 보내지는 엑셀메일 정보입니다.
        엑셀내용을 잘 첨부하세요.
        """

        msg.attach(MIMEText(text))

        s=smtplib.SMTP(smtp_name, smtp_port)
        s.starttls()
        s.login(send_email, send_pwd)
        s.sendmail(send_email, recv_email, msg.as_string())
        s.quit()

    except:
        print("에러:", recv_email_value)
